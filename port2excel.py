"""
Name:       port2excel.py
Objective:  take a list of numbers. port each num into cells in excel
Params:     list of number
Return:      None
Author:      Project Echo Telion [echo.telion@gmail.com]

"""

from openpyxl import load_workbook
from parsing_module import get_data
from parsing_module import get_date

def port_num(num_list):
    workbook = load_workbook('project_damacai.xlsx')
    worksheet = workbook.create_sheet(0)
    worksheet.title = get_date()       #get date range selected and use it as title for newly created worksheet

    init_row = 2
    init_col = 1

# port data into excel worksheet
    for each_element in num_list:
        for each_num in each_element:  
            worksheet.cell(row = init_row, column = init_col).value = each_num
            init_row += 1
            
        init_col += 1
        init_row = 2

# label each column in worksheet
    worksheet_headings = ["NUMBER","TOTAL","TOP 3","1ST","2ND","3RD","STARTER","CONSO"]
    col_counter = 1
    for each in worksheet_headings:
        worksheet.cell(row = 1, column = col_counter).value = each
        col_counter += 1
        
    workbook.save('project_damacai.xlsx')
